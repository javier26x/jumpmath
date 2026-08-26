"""Lectura tolerante de planillas (.xlsx / .csv).

Los colegios no comparten un formato único: cambian el orden de las columnas,
el texto de los encabezados y la fila donde empieza la tabla (suele haber un
logo y un par de líneas de título arriba). En vez de fijar coordenadas, aquí
se busca la fila de encabezado que mejor casa con un conjunto de sinónimos
por columna.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from typing import Any

from ..errores import ErrorParseo
from ..normalizacion import clave, es_vacio, limpiar

#: Filas iniciales en las que se busca el encabezado antes de rendirse.
MAX_FILAS_ENCABEZADO = 25


@dataclass(slots=True)
class Columna:
    """Una columna esperada y las formas en que puede venir escrita."""

    nombre: str
    sinonimos: tuple[str, ...]
    requerida: bool = True

    def casa(self, celda: str) -> bool:
        c = clave(celda)
        if not c:
            return False
        return any(c == clave(s) or clave(s) in c for s in self.sinonimos)


@dataclass(slots=True)
class Tabla:
    """Tabla localizada dentro de una hoja: encabezado resuelto + filas."""

    fuente: str
    hoja: str
    indice_col: dict[str, int]
    filas: list[list[Any]]

    def valor(self, fila: Sequence[Any], columna: str) -> Any:
        """Celda de `fila` bajo la columna lógica `columna` (o `None`)."""
        i = self.indice_col.get(columna)
        if i is None or i >= len(fila):
            return None
        return fila[i]

    def registros(self) -> Iterator[dict[str, Any]]:
        """Filas como diccionarios, omitiendo las completamente vacías."""
        for fila in self.filas:
            if all(es_vacio(c) for c in fila):
                continue
            yield {nombre: self.valor(fila, nombre) for nombre in self.indice_col}


def _hojas(datos: bytes, nombre_archivo: str) -> list[tuple[str, list[list[Any]]]]:
    """Devuelve `[(hoja, filas)]` para .xlsx/.xlsm o un CSV de una sola hoja."""
    if nombre_archivo.lower().endswith((".csv", ".tsv", ".txt")):
        texto = datos.decode("utf-8-sig", errors="replace")
        dialecto = (
            csv.Sniffer().sniff(texto[:4096], delimiters=",;\t") if texto.strip() else csv.excel
        )
        filas = [list(f) for f in csv.reader(io.StringIO(texto), dialecto)]
        return [("csv", filas)]

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependencia de despliegue
        raise ErrorParseo(nombre_archivo, f"falta openpyxl para leer planillas: {exc}") from exc

    try:
        libro = load_workbook(io.BytesIO(datos), data_only=True, read_only=True)
    except Exception as exc:
        raise ErrorParseo(nombre_archivo, f"no es un .xlsx legible: {exc}") from exc

    salida = []
    for hoja in libro.worksheets:
        salida.append((hoja.title, [list(f) for f in hoja.iter_rows(values_only=True)]))
    libro.close()
    return salida


def localizar_tabla(
    datos: bytes,
    nombre_archivo: str,
    columnas: Sequence[Columna],
    hoja_preferida: str | None = None,
) -> Tabla:
    """Encuentra la hoja y la fila de encabezado que mejor casan con `columnas`.

    Puntúa cada fila candidata por cuántas columnas requeridas resuelve y se
    queda con la mejor de todas las hojas. Falla con el detalle de lo que no
    encontró, que es lo que necesita quien tenga que corregir la planilla.
    """
    requeridas = [c for c in columnas if c.requerida]
    mejor: Tabla | None = None
    mejor_puntaje = -1

    for nombre_hoja, filas in _hojas(datos, nombre_archivo):
        if hoja_preferida and clave(hoja_preferida) not in clave(nombre_hoja):
            continue
        for i, fila in enumerate(filas[:MAX_FILAS_ENCABEZADO]):
            indice: dict[str, int] = {}
            for col in columnas:
                for j, celda in enumerate(fila):
                    if j in indice.values():
                        continue
                    if col.casa(limpiar(celda)):
                        indice[col.nombre] = j
                        break
            puntaje = sum(1 for c in requeridas if c.nombre in indice)
            if puntaje > mejor_puntaje:
                mejor_puntaje = puntaje
                mejor = Tabla(nombre_archivo, nombre_hoja, indice, filas[i + 1 :])

    if mejor is None or mejor_puntaje < len(requeridas):
        faltan = [c.nombre for c in requeridas if not mejor or c.nombre not in mejor.indice_col]
        raise ErrorParseo(
            nombre_archivo,
            "no se encontró la tabla esperada; faltan columnas: "
            + ", ".join(faltan)
            + ". Sinónimos aceptados: "
            + "; ".join(
                f"{c.nombre}={c.sinonimos}" for c in requeridas if c.nombre in faltan
            ),
        )
    return mejor
