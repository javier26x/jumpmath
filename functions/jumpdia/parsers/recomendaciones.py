"""Parser de «Recomendaciones por indicador · 4° Básico» (.xlsx).

Alimenta `recs[].{base, plus}` y —lo más importante— el mapeo
**indicador → unidad JUMP** (Tomo 4.1 / 4.2) del que dependen `recs[].units`
y `recs[].estado`.

Sobre el formato real: el encabezado ocupa **dos filas**. «Unidad JUMP Math»
abarca dos columnas y sólo la fila de abajo dice cuál es «Tomo 4.1» y cuál
«Tomo 4.2»; cada celda trae `U4`, o un guion si esa unidad no cubre el
indicador. Un indicador puede caer en los dos tomos a la vez.

La planilla identifica cada fila por **N° de pregunta**, que es la clave que se
usa para cruzar con el informe oficial: emparejar por el texto del indicador
sería frágil ante cualquier diferencia de puntuación o tilde.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

from ..catalogo import POR_CLAVE, UnidadJump
from ..errores import ErrorParseo
from ..normalizacion import a_float, clave, es_vacio, limpiar, texto_largo
from .planilla import Columna, localizar_tabla

FUENTE = "recomendaciones por indicador"

#: "Tomo 4.1 U2", "4.1·U2", "4.1 - Unidad 2" … → ("4.1", "U2")
_RE_UNIDAD = re.compile(
    r"(?:tomo\s*)?(4\s*[.,]\s*[12])\s*[·\-–—/ ]*\s*(?:u(?:nidad)?\s*)?([1-8])\b", re.I
)
#: Cabecera de una columna dedicada a un tomo: "Tomo 4.2".
_RE_COL_TOMO = re.compile(r"tomo\s*4\s*[.,]\s*([12])\s*$", re.I)
#: Contenido de esas columnas: "U4", "U4-U5", "---".
_RE_U = re.compile(r"u\s*([1-8])\b", re.I)

_COLUMNAS = (
    Columna("ind", ("indicador", "indicador de evaluacion", "descripcion del indicador")),
    Columna("base", ("sugerencias", "recomendacion", "sugerencia", "orientaciones")),
    Columna("q", ("n de pregunta", "pregunta", "n pregunta", "item"), requerida=False),
    Columna("oa", ("oa", "oa evaluado", "objetivo de aprendizaje"), requerida=False),
    Columna(
        "unidades", ("unidad jump", "unidades jump", "tomo y unidad", "unidad"), requerida=False
    ),
    Columna(
        "plus",
        ("analisis", "analisis adicional", "comentario", "nota jump", "plus"),
        requerida=False,
    ),
)


@dataclass(slots=True)
class ResultadoRecomendaciones:
    """Textos y mapeo a unidades JUMP, indexados por N° de pregunta.

    Se conserva también el índice por indicador, para planillas que no traen
    el número de pregunta.
    """

    base: dict[int, str] = field(default_factory=dict)
    plus: dict[int, str] = field(default_factory=dict)
    unidades: dict[int, list[UnidadJump]] = field(default_factory=dict)
    base_por_indicador: dict[str, str] = field(default_factory=dict)
    plus_por_indicador: dict[str, str] = field(default_factory=dict)
    unidades_por_indicador: dict[str, list[UnidadJump]] = field(default_factory=dict)
    avisos: list[str] = field(default_factory=list)


def extraer_unidades(texto: object) -> list[UnidadJump]:
    """Extrae las unidades JUMP citadas en una celda, sin repetir y en orden."""
    encontradas: list[UnidadJump] = []
    for tomo, numero in _RE_UNIDAD.findall(limpiar(texto)):
        unidad = POR_CLAVE.get(f"{tomo.replace(' ', '').replace(',', '.')}/U{numero}")
        if unidad is not None and unidad not in encontradas:
            encontradas.append(unidad)
    return encontradas


def _columnas_por_tomo(datos: bytes, nombre_archivo: str) -> dict[str, int]:
    """Índice de la columna dedicada a cada tomo, si el encabezado las separa.

    Se busca en el libro completo la celda «Tomo 4.1» / «Tomo 4.2» usada como
    subencabezado. `localizar_tabla` resuelve una sola fila de cabecera y por
    eso no puede verlas.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependencia de despliegue
        raise ErrorParseo(
            f"{FUENTE} · {nombre_archivo}", f"falta openpyxl para leer planillas: {exc}"
        ) from exc

    columnas: dict[str, int] = {}
    libro = load_workbook(io.BytesIO(datos), data_only=True, read_only=True)
    for hoja in libro.worksheets:
        for fila in hoja.iter_rows(max_row=12, values_only=True):
            for j, celda in enumerate(fila):
                if m := _RE_COL_TOMO.match(limpiar(celda)):
                    columnas.setdefault(f"4.{m.group(1)}", j)
        if len(columnas) == 2:
            break
    libro.close()
    return columnas


def _unidades_de_fila(fila: list, columnas_tomo: dict[str, int]) -> list[UnidadJump]:
    """Unidades JUMP de una fila leídas de sus columnas por tomo."""
    unidades: list[UnidadJump] = []
    for tomo, j in sorted(columnas_tomo.items()):
        celda = limpiar(fila[j]) if j < len(fila) else ""
        for numero in _RE_U.findall(celda):
            if (unidad := POR_CLAVE.get(f"{tomo}/U{numero}")) and unidad not in unidades:
                unidades.append(unidad)
    return unidades


def parsear_recomendaciones(datos: bytes, nombre_archivo: str) -> ResultadoRecomendaciones:
    """Lee la planilla de recomendaciones y su mapeo a unidades JUMP."""
    tabla = localizar_tabla(datos, nombre_archivo, _COLUMNAS)
    columnas_tomo = _columnas_por_tomo(datos, nombre_archivo)
    resultado = ResultadoRecomendaciones()

    for fila in tabla.filas:
        indicador = limpiar(tabla.valor(fila, "ind"))
        if not indicador or es_vacio(indicador):
            continue

        unidades = _unidades_de_fila(fila, columnas_tomo)
        if not unidades:
            unidades = extraer_unidades(tabla.valor(fila, "unidades"))
        if not unidades:
            resultado.avisos.append(
                f"«{indicador[:55]}…»: no se reconoció ninguna unidad JUMP."
            )

        base = texto_largo(tabla.valor(fila, "base"))
        plus = texto_largo(tabla.valor(fila, "plus"))
        llave_ind = clave(indicador)
        resultado.base_por_indicador[llave_ind] = base
        resultado.plus_por_indicador[llave_ind] = plus
        resultado.unidades_por_indicador[llave_ind] = unidades

        numero = a_float(tabla.valor(fila, "q"))
        if numero is not None:
            resultado.base[int(numero)] = base
            resultado.plus[int(numero)] = plus
            resultado.unidades[int(numero)] = unidades

    if not resultado.base_por_indicador:
        raise ErrorParseo(FUENTE, "la planilla no contiene indicadores")
    if not columnas_tomo:
        resultado.avisos.append(
            "el encabezado no separa «Tomo 4.1» y «Tomo 4.2»: las unidades se "
            "leyeron de una sola columna."
        )
    return resultado
