"""Parser del «Seguimiento de evaluaciones JUMP» (controles y pruebas, .xlsx).

Alimenta `D.coverage[]`. Regla central de la guía (§4): sólo se marca
`status:'res'` una unidad con control o prueba **registrada**. La ausencia de
registro es `status:'none'` y **no** significa "no trabajada" — puede estar
trabajada sin registrar o con la prueba pendiente.

Sobre el formato real: el libro trae **una hoja por evaluación** («Control 1
U1», «Prueba U2», «Prueba U7 - 4.2»…), no una tabla de unidades con su
promedio. Cada hoja es la corrección ítem a ítem —un 1 o un 0 por estudiante y
pregunta— así que el porcentaje de la unidad hay que **calcularlo**. Una unidad
puede tener varias evaluaciones, y hojas preparadas pero aún sin aplicar, que
son precisamente las que deben quedar como "sin registro".
"""

from __future__ import annotations

import io
import re
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any

from ..catalogo import UNIDADES_JUMP, unidad_por_nombre
from ..errores import ErrorParseo
from ..normalizacion import a_float, clave, limpiar
from .planilla import Columna, localizar_tabla
from .recomendaciones import extraer_unidades

FUENTE = "seguimiento de evaluaciones JUMP"

#: «Control 1 Unidad 1: Series» · «Prueba Unidad 6: Área y volumen»
_RE_TITULO = re.compile(r"unidad\s*(\d)\s*[:.\-]\s*(.+)", re.I)

#: Encabezados de la fila de preguntas que no son preguntas.
_NO_PREGUNTA = ("total", "logro", "porcentaje", "puntaje", "nota", "estudiante", "n")

#: Comienzos de nombre que delatan una fila de resumen, no un estudiante.
_FILAS_RESUMEN = ("total", "promedio", "suma", "curso", "logro", "porcentaje")

#: Formato antiguo: una fila por unidad con su promedio ya calculado.
_COLUMNAS_RESUMEN = (
    Columna("unidad", ("unidad", "unidad jump", "tomo y unidad", "tomo")),
    Columna("pct", ("promedio", "logro", "porcentaje", "% logro", "resultado")),
    Columna("estado", ("estado", "aplicado", "rendido", "registrado"), requerida=False),
)


@dataclass(slots=True)
class ResultadoSeguimiento:
    coverage: list[dict[str, Any]] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)


@dataclass(slots=True)
class _Evaluacion:
    """Una hoja de corrección ya resumida."""

    hoja: str
    aciertos: float
    respuestas: int

    @property
    def pct(self) -> float:
        return 100.0 * self.aciertos / self.respuestas


def _unidad_de_hoja(titulo: str, nombre_hoja: str):
    """Unidad JUMP a la que corresponde una hoja, por su título o su nombre.

    El nombre del nivel («Series», «Figuras») es lo que distingue el tomo, así
    que se prefiere el título de la hoja, que lo lleva. El nombre de la
    pestaña sólo trae el número y un sufijo inconsistente (`U1 - 2`, `U7 - 4.2`).
    """
    m = _RE_TITULO.search(limpiar(titulo))
    if m and (unidad := unidad_por_nombre(int(m.group(1)), m.group(2))):
        return unidad
    unidades = extraer_unidades(nombre_hoja)
    return unidades[0] if unidades else None


def _resumir_hoja(filas: list[list[Any]]) -> tuple[float, int]:
    """Aciertos y respuestas registradas de una hoja de corrección.

    Devuelve `(0, 0)` para una hoja preparada pero sin aplicar. El denominador
    cuenta sólo las celdas con marca: un estudiante ausente no debe bajar el
    logro del curso como si hubiera fallado todo.

    Sólo cuentan las filas con nombre de estudiante. La hoja lleva además, bajo
    la nómina, un bloque de logro por indicador y una fila de totales; esa
    última viene con ceros de fórmula y, contada como si fuera un alumno,
    arrastraría hacia abajo el resultado de toda la unidad.
    """
    encabezado, columnas, col_nombre = None, [], None
    for i, fila in enumerate(filas):
        etiquetas = [clave(c) for c in fila]
        if "estudiante" not in etiquetas:
            continue
        preguntas = [
            j
            for j, e in enumerate(etiquetas)
            if a_float(e) is not None and not any(t in e for t in _NO_PREGUNTA)
        ]
        if preguntas:
            encabezado, columnas = i, preguntas
            col_nombre = etiquetas.index("estudiante")
            break

    if encabezado is None:
        return 0.0, 0

    aciertos, respuestas = 0.0, 0
    for fila in filas[encabezado + 1 :]:
        nombre = clave(fila[col_nombre]) if col_nombre < len(fila) else ""
        # «Total por pregunta» cierra la nómina con la suma de cada columna y
        # tiene exactamente la forma de un estudiante: hay que descartarla por
        # el nombre, o el curso sumaría un alumno con 259 aciertos de 10.
        if not nombre or nombre.startswith(_FILAS_RESUMEN):
            continue
        marcas = [a_float(fila[j]) if j < len(fila) else None for j in columnas]
        marcas = [m for m in marcas if m is not None]
        # El bloque de indicadores repite la forma de una fila de estudiante
        # pero deja vacías las preguntas.
        if not marcas:
            continue
        aciertos += sum(min(max(m, 0.0), 1.0) for m in marcas)
        respuestas += len(marcas)
    return aciertos, respuestas


def _por_hojas(datos: bytes, resultado: ResultadoSeguimiento):
    """Recorre el libro tratando cada hoja como una evaluación."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependencia de despliegue
        raise ErrorParseo(FUENTE, f"falta openpyxl para leer planillas: {exc}") from exc

    try:
        libro = load_workbook(io.BytesIO(datos), data_only=True, read_only=True)
    except Exception as exc:
        raise ErrorParseo(FUENTE, f"no es un .xlsx legible: {exc}") from exc

    evaluaciones: dict[str, list[_Evaluacion]] = defaultdict(list)
    for hoja in libro.worksheets:
        filas = [list(f) for f in hoja.iter_rows(values_only=True)]
        titulo = next((limpiar(c) for f in filas[:3] for c in f if limpiar(c)), "")
        unidad = _unidad_de_hoja(titulo, hoja.title)
        if unidad is None:
            continue

        aciertos, respuestas = _resumir_hoja(filas)
        if respuestas == 0:
            resultado.avisos.append(
                f"«{hoja.title}» ({unidad.clave} {unidad.label}) no tiene resultados "
                "registrados: la unidad queda sin registro, que no equivale a no trabajada."
            )
            continue
        evaluaciones[unidad.clave].append(_Evaluacion(hoja.title, aciertos, respuestas))

    libro.close()
    return evaluaciones


def parsear_seguimiento(datos: bytes, nombre_archivo: str) -> ResultadoSeguimiento:
    """Construye la cobertura JUMP en el orden curricular del catálogo."""
    resultado = ResultadoSeguimiento()
    evaluaciones = _por_hojas(datos, resultado)

    if not evaluaciones:
        # Ninguna hoja parecía una corrección: puede ser el formato resumido,
        # con una fila por unidad y el promedio ya calculado.
        evaluaciones = _formato_resumen(datos, nombre_archivo)

    for unidad in UNIDADES_JUMP:
        aplicadas = evaluaciones.get(unidad.clave, [])
        if len(aplicadas) > 1:
            resultado.avisos.append(
                f"{unidad.clave} {unidad.label}: {len(aplicadas)} evaluaciones registradas "
                f"({', '.join(e.hoja for e in aplicadas)}); se informa su promedio."
            )
        pct = sum(e.pct for e in aplicadas) / len(aplicadas) if aplicadas else None
        resultado.coverage.append(
            {
                "tomo": unidad.tomo,
                "u": unidad.u,
                "label": unidad.label,
                "status": "res" if pct is not None else "none",
                "pct": round(pct, 1) if pct is not None else None,
            }
        )

    if not any(c["status"] == "res" for c in resultado.coverage):
        resultado.avisos.append(
            "no se registró ningún control o prueba: toda la cobertura queda "
            "en 'none' (sin registro), que no equivale a 'no trabajada'."
        )
    return resultado


def _formato_resumen(datos: bytes, nombre_archivo: str):
    """Lee el formato de una fila por unidad con su promedio ya calculado."""
    evaluaciones: dict[str, list[_Evaluacion]] = defaultdict(list)
    try:
        tabla = localizar_tabla(datos, nombre_archivo, _COLUMNAS_RESUMEN)
    except ErrorParseo:
        return evaluaciones

    from ..normalizacion import a_pct, es_vacio

    for registro in tabla.registros():
        unidades = extraer_unidades(registro.get("unidad"))
        estado = clave(registro.get("estado"))
        if not unidades or estado in ("pendiente", "no aplicado", "no rendido", "no"):
            continue
        pct = a_pct(registro.get("pct"))
        if pct is None or es_vacio(registro.get("pct")):
            continue
        for unidad in unidades:
            evaluaciones[unidad.clave].append(_Evaluacion(unidad.clave, pct, 100))
    return evaluaciones
