"""Genera archivos de entrada sintéticos que reproducen el informe 4° A.

Los archivos reales del colegio no están en el repositorio (traen datos
personales de estudiantes). Estos fixtures replican su **estructura** a partir
del `D` publicado, de modo que los tests puedan recorrer el pipeline completo
—PDF y planillas incluidos— y comprobar que se vuelve a obtener ese mismo `D`.

Uso:  python tests/generar_fixtures.py
"""

from __future__ import annotations

import json
import pathlib
import sys

RAIZ = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ / "functions"))

from jumpdia.catalogo import EJE_CLAVES, EJES  # noqa: E402

FIXTURES = RAIZ / "tests" / "fixtures"
D = json.loads((FIXTURES / "D_4A_bulnes.json").read_text(encoding="utf-8"))

#: Categorías de la distribución en el orden en que las imprime el DIA.
COLUMNAS_DIST = ("A", "B", "C", "D", "RC", "RPC", "RI", "N")


def aciertos_por_eje(alumno: dict) -> list[float]:
    """Reconstruye los aciertos crudos de cada eje desde su % redondeado.

    El informe publica porcentajes enteros; los aciertos son múltiplos de 0,5
    (medio punto por respuesta parcialmente correcta en ítems de desarrollo).
    """
    salida = []
    for llave, total in zip(EJE_CLAVES, D["ejeQC"], strict=True):
        salida.append(round(alumno[llave] / 100 * total * 2) / 2)
    return salida


# --- Informe oficial DIA (PDF) -------------------------------------------


def escribir_dia_pdf(destino: pathlib.Path, questions=None, meta=None) -> None:
    """Escribe un PDF con la disposición del informe oficial real.

    Importa reproducir la **forma**, no sólo los datos: la Tabla 1 no tiene
    rejilla —es texto en columnas que el parser localiza por coordenadas— y la
    alternativa correcta viene destacada en negrita. Un fixture con una tabla
    de bordes dibujados pasaría los tests sin ejercitar nada de eso.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    questions = questions if questions is not None else D["questions"]
    meta = meta if meta is not None else D["meta"]

    normal = getSampleStyleSheet()["Normal"]
    celda = ParagraphStyle("celda", parent=normal, fontSize=6.5, leading=8)
    pct = ParagraphStyle("pct", parent=celda, fontName="Helvetica")

    flujo = [
        Paragraph("· Diagnóstico Integral de Aprendizajes ·", normal),
        Paragraph("Informe de Resultados 2026", normal),
        Paragraph("Prueba de Matemática", normal),
        Paragraph(f"Establecimiento:{meta['colegio']}", normal),
        Paragraph(f"RBD:{meta['rbd']}", normal),
        Paragraph(f"Nombre director o directora: {meta['director']}", normal),
        Paragraph(f"Nombre docente de la asignatura:{meta['docente']}", normal),
        Paragraph(f"Curso: {meta['curso'].replace(chr(176), '')}", normal),
        Paragraph(
            f"Cantidad de estudiantes que considera este informe: {meta['n']}", normal
        ),
        Paragraph(
            f"Fecha y hora de generación de este informe: {meta['fecha']} 12:31:18", normal
        ),
        Paragraph("Monitoreo Intermedio", normal),
        Spacer(1, 12),
        Paragraph("Tabla 1. Resultados del curso en cada pregunta de la prueba", normal),
        Spacer(1, 6),
    ]

    encabezado = [
        Paragraph(f"<b>{h}</b>", celda)
        for h in (
            "N° pregunta", "N° OA", "Eje temático", "Habilidad",
            "Indicador de evaluación", "% respuestas",
        )
    ]
    filas = [encabezado]
    for q in questions:
        filas.append(
            [
                Paragraph(str(q["q"]), celda),
                Paragraph(str(q["oa"]), celda),
                Paragraph(q["eje"], celda),
                Paragraph(q["hab"], celda),
                Paragraph(q["ind"], celda),
                Paragraph(_lineas_distribucion(q), pct),
            ]
        )

    tabla = Table(filas, colWidths=[52, 40, 74, 78, 176, 72], repeatRows=1)
    tabla.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    flujo += [tabla, Spacer(1, 10), Paragraph("Preguntas guía", normal)]

    SimpleDocTemplate(
        str(destino), pagesize=A4, topMargin=26, bottomMargin=26,
        leftMargin=30, rightMargin=30,
    ).build(flujo)


def _lineas_distribucion(q: dict) -> str:
    """Columna «% respuestas», con la alternativa correcta en negrita.

    En un ítem de alternativas la clave es la única marca de cuál es la
    respuesta correcta; se destaca la de mayor porcentaje, que es la que el
    informe publicado usó como logro del ítem.
    """
    dist = {k: v for k, v in q["dist"].items() if k != "clave"}
    clave_correcta = q["dist"].get("clave")
    if clave_correcta is None and q["tipo"] == "alternativas":
        opciones = {k: v for k, v in dist.items() if k != "N"}
        clave_correcta = max(opciones, key=opciones.get) if opciones else None

    lineas = []
    for etiqueta in ("RC", "RPC", "RI", "A", "B", "C", "D", "E", "N"):
        if etiqueta not in dist:
            continue
        texto = f"{etiqueta}: {dist[etiqueta]:.2f}%"
        lineas.append(f"<b>{texto}</b>" if etiqueta == clave_correcta else texto)
    return "<br/>".join(lineas)


# --- Planillas ------------------------------------------------------------


def _libro(hoja: str, filas: list[list], destino: pathlib.Path, preambulo: int = 2) -> None:
    """Escribe un .xlsx con `preambulo` filas de título antes de la tabla."""
    from openpyxl import Workbook

    libro = Workbook()
    ws = libro.active
    ws.title = hoja
    for _ in range(preambulo):
        ws.append(["Colegio Manuel Bulnes Prieto"])
    for fila in filas:
        ws.append(fila)
    libro.save(destino)


def escribir_estudiantes(destino: pathlib.Path) -> None:
    filas = [["Estudiante", *EJES, "Puntaje obtenido", "Nivel de aprendizaje"]]
    for alumno in D["students"]:
        aciertos = aciertos_por_eje(alumno)
        filas.append(
            [
                alumno["n"],
                *[alumno[k] for k in EJE_CLAVES],
                sum(aciertos),
                {1: "Nivel I", 2: "Nivel II", 3: "Nivel III"}[alumno["lv"]],
            ]
        )
    _libro("Resultados", filas, destino)


def _partes_unidades(units: str) -> list[str]:
    """Parte `"4.1 · U2 Nombre ✓ 85% · 4.2 · U3 Otro ✗"` en unidades sueltas."""
    tokens, actual = [], []
    for token in units.split(" · "):
        actual.append(token)
        if len(actual) == 2:
            tokens.append(" · ".join(actual))
            actual = []
    return tokens


def escribir_recomendaciones(destino: pathlib.Path) -> None:
    """Planilla de recomendaciones con el encabezado real, de dos filas.

    «Unidad JUMP Math» abarca dos columnas y sólo la fila de abajo dice cuál
    es cada tomo; el cruce con el informe se hace por N° de pregunta.
    """
    from openpyxl import Workbook

    libro = Workbook()
    ws = libro.active
    ws.title = "4° A"
    ws.append(["Indicadores que requieren refuerzo - 4° Básico"])
    ws.append(
        ["N° de pregunta", "OA evaluado", "Indicador", "% de logro\n4° A",
         "Unidad JUMP Math", None, "Sugerencias metodológicas y/o didácticas",
         "Análisis adicional"]
    )
    ws.append([None, None, None, None, "Tomo 4.1", "Tomo 4.2", None, None])

    for rec in D["recs"]:
        unidades = {}
        for parte in _partes_unidades(rec["units"]):
            tomo, resto = parte.split(" · ", 1)
            unidades[tomo.strip()] = resto.split(" ", 1)[0]
        ws.append(
            [
                rec["q"], rec["oa"], rec["ind"], None,
                unidades.get("4.1", "-"), unidades.get("4.2", "-"),
                rec["base"], rec["plus"],
            ]
        )
    libro.save(destino)


#: Preguntas por evaluación en el fixture. Con 26 estudiantes da 1 300 celdas,
#: de modo que cualquier porcentaje entero se alcanza con un número exacto de
#: aciertos y el fixture reproduce la cobertura publicada al decimal.
PREGUNTAS_POR_CONTROL = 50


def escribir_seguimiento(destino: pathlib.Path) -> None:
    """Libro de seguimiento con el formato real: una hoja por evaluación.

    Cada hoja es la corrección ítem a ítem, así que el porcentaje de la unidad
    hay que calcularlo. Se incluye a propósito lo que trae el archivo del
    colegio y confunde a un parser ingenuo: la fila «Total por pregunta» al pie
    de la nómina, y hojas preparadas pero sin aplicar, que deben quedar como
    unidades sin registro y no como un 0 %.
    """
    from openpyxl import Workbook

    libro = Workbook()
    libro.remove(libro.active)
    nombres = [alumno["n"] for alumno in D["students"]]

    for cobertura in D["coverage"]:
        titulo = f"{'Prueba' if cobertura['status'] == 'res' else 'Control'} Unidad "
        titulo += f"{cobertura['u'][1:]}: {cobertura['label']}"
        ws = libro.create_sheet(f"{cobertura['tomo']} {cobertura['u']}"[:31])
        ws.append([titulo])
        ws.append([])
        ws.append([None, None, "Preguntas"])
        ws.append(
            ["N°", "Estudiante", *range(1, PREGUNTAS_POR_CONTROL + 1), "Total", "% logro"]
        )

        if cobertura["status"] != "res":
            # Hoja lista pero sin aplicar: la nómina va vacía.
            for i, nombre in enumerate(nombres, 1):
                ws.append([i, nombre])
            ws.append([None, "Total por pregunta"])
            continue

        celdas = len(nombres) * PREGUNTAS_POR_CONTROL
        restantes = round(cobertura["pct"] / 100 * celdas)
        for i, nombre in enumerate(nombres, 1):
            aciertos = min(restantes, PREGUNTAS_POR_CONTROL)
            restantes -= aciertos
            marcas = [1] * aciertos + [0] * (PREGUNTAS_POR_CONTROL - aciertos)
            ws.append([i, nombre, *marcas, sum(marcas), sum(marcas) / PREGUNTAS_POR_CONTROL])
        ws.append([None, "Total por pregunta", *([0] * PREGUNTAS_POR_CONTROL)])

    libro.save(destino)


def main() -> None:
    FIXTURES.mkdir(parents=True, exist_ok=True)
    escribir_dia_pdf(FIXTURES / "dia_oficial_4A.pdf")
    escribir_estudiantes(FIXTURES / "estudiantes_4A.xlsx")
    escribir_recomendaciones(FIXTURES / "recomendaciones_4B.xlsx")
    escribir_seguimiento(FIXTURES / "seguimiento_jump_4A.xlsx")
    for archivo in sorted(FIXTURES.iterdir()):
        print(f"  {archivo.name:32s} {archivo.stat().st_size:>8,} bytes")


if __name__ == "__main__":
    main()
